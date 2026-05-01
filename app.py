"""
app.py
======
Flask backend for the IATA Fuel Monitor — tables downloader.

Routes
------
  GET  /          → renders the UI
  POST /download  → scrapes tables → Excel workbook (appends sheet)
  GET  /status    → JSON: sheet count + last update time
  POST /reset     → deletes workbook so next download starts fresh
"""

import io
import traceback
from pathlib import Path

from flask import Flask, render_template, send_file, jsonify, Response

from scrape import scrape_tables
from excel  import build_or_update, WORKBOOK_PATH

app = Flask(__name__)


@app.route("/")
def index():
    sheets = _sheet_info()
    return render_template(
        "index.html",
        sheet_count  = sheets["count"],
        last_updated = sheets["last_updated"],
    )


@app.route("/download", methods=["POST"])
def download():
    try:
        t1_rows, t2_rows = scrape_tables()

        existing = WORKBOOK_PATH.read_bytes() if WORKBOOK_PATH.exists() else None
        wb_bytes, is_dup, sheet_name = build_or_update(t1_rows, t2_rows, existing)

        if not is_dup:
            WORKBOOK_PATH.parent.mkdir(parents=True, exist_ok=True)
            WORKBOOK_PATH.write_bytes(wb_bytes)

        response = Response(
            wb_bytes,
            mimetype=(
                "application/vnd.openxmlformats-"
                "officedocument.spreadsheetml.sheet"
            ),
            headers={
                "Content-Disposition":
                    'attachment; filename="IATA_Fuel_Tables.xlsx"',
            },
        )
        response.headers["X-Is-Duplicate"] = "1" if is_dup else "0"
        response.headers["X-Sheet-Name"]   = sheet_name
        response.headers["Access-Control-Expose-Headers"] = (
            "X-Is-Duplicate, X-Sheet-Name"
        )
        return response

    except Exception as exc:
        tb = traceback.format_exc()
        app.logger.error(f"Download failed:\n{tb}")
        return jsonify({"error": str(exc), "detail": tb}), 500


@app.route("/status")
def status():
    return jsonify(_sheet_info())


@app.route("/reset", methods=["POST"])
def reset():
    try:
        if WORKBOOK_PATH.exists():
            WORKBOOK_PATH.unlink()
            return jsonify({"ok": True,
                            "message": "Workbook deleted. Next download starts fresh."})
        return jsonify({"ok": True, "message": "No workbook found — already clean."})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500


def _sheet_info() -> dict:
    if not WORKBOOK_PATH.exists():
        return {"count": 0, "last_updated": None, "sheets": []}
    try:
        from openpyxl import load_workbook
        from excel import SHEET_CONSOL_T1, SHEET_CONSOL_T2
        wb        = load_workbook(WORKBOOK_PATH, read_only=True)
        names     = list(wb.sheetnames)
        wb.close()
        snapshots = [s for s in names if s not in (SHEET_CONSOL_T1, SHEET_CONSOL_T2)]
        return {
            "count":        len(snapshots),
            "last_updated": snapshots[-1] if snapshots else None,
            "sheets":       names,
        }
    except Exception:
        return {"count": 0, "last_updated": None, "sheets": []}


if __name__ == "__main__":
    WORKBOOK_PATH.parent.mkdir(parents=True, exist_ok=True)
    app.run(debug=True, port=5050)