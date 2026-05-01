"""
excel.py
========
Builds / updates the Excel workbook.

Workbook structure
------------------
  Sheet "Consolidated T1"  ← ALWAYS first; accumulates Table 1 across all downloads
  Sheet "Consolidated T2"  ← ALWAYS second; accumulates Table 2 across all downloads
  Sheet "YYYY-MM-DD HH.MM" ← snapshot of this download (both tables, existing behaviour)
  Sheet "YYYY-MM-DD HH.MM" ← next download snapshot …

Consolidated sheets layout
---------------------------
  Column A  : "Extraction Date" — the datetime when that block was downloaded.
              Filled only on the first data row of each block; blank for the rest.
  Columns B+ : exactly the same columns as the original table (headers repeat
              at the top of every new block so each block is self-contained).

  First download:
    Row 1  : header  (Extraction Date | col1 | col2 | …)
    Row 2  : data row 1   ← "2026-04-30 16.45" in col A
    Row 3  : data row 2   ← blank in col A
    …

  Second download (data changed):
    (blank spacer row)
    Row N  : header  (Extraction Date | col1 | col2 | …)
    Row N+1: data row 1   ← "2026-05-06 16.12" in col A
    …

Duplicate detection
-------------------
  If the scraped data is identical to the last download, no new snapshot sheet
  is created and the consolidated sheets are NOT appended to (no duplicate blocks).
"""

import hashlib
import json
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Colours ───────────────────────────────────────────────────────────────────
C_NAVY      = "1F3864"
C_GREY      = "D9D9D9"
C_BLUE      = "2E4DA7"
C_PALE_BG   = "C5D1EB"
C_IDX_DATA  = "DCE6F1"
C_ALT_ROW   = "F2F2F2"
C_DATE_BG   = "E8F0FE"   # pale blue tint for Extraction Date column
C_WHITE     = "FFFFFF"
C_BLACK     = "000000"
C_BORDER    = "BFBFBF"

# On Render use the persistent disk at /data; locally use workbooks/
import os as _os
_wb_dir = _os.environ.get("WORKBOOK_DIR", "workbooks")
WORKBOOK_PATH   = Path(_wb_dir) / "iata_fuel_tables.xlsx"
SHEET_CONSOL_T1 = "Consolidated T1"
SHEET_CONSOL_T2 = "Consolidated T2"


# ── Style helpers ─────────────────────────────────────────────────────────────
def _fill(hex_c):
    return PatternFill("solid", fgColor=hex_c)

def _font(bold=False, color=C_BLACK, size=10):
    return Font(bold=bold, color=color, size=size, name="Calibri")

def _align(h="left", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def _border():
    s = Side(style="thin", color=C_BORDER)
    return Border(left=s, right=s, top=s, bottom=s)

def _cell(ws, r, c, value="", *, bold=False, bg=None, fg=C_BLACK,
          h_align="left", wrap=False):
    cell           = ws.cell(row=r, column=c, value=value)
    cell.font      = _font(bold=bold, color=fg)
    cell.alignment = _align(h=h_align, wrap=wrap)
    cell.border    = _border()
    if bg:
        cell.fill = _fill(bg)
    return cell

def _hdr(ws, r, c, text, bg, fg=C_WHITE, wrap=True):
    _cell(ws, r, c, text, bold=True, bg=bg, fg=fg, h_align="center", wrap=wrap)


# ── Snapshot: Table 1 (cols 1-9) ─────────────────────────────────────────────
def _write_table1(ws, rows, start_row, col_offset=0):
    """
    Write Table 1 starting at start_row.
    col_offset: shift all columns right by this many (used in snapshot sheet = 0).
    Returns last written row.
    """
    r  = start_row
    c0 = col_offset  # 0 for snapshot, 1 for consolidated (col A = extraction date)

    # Super-header row 1
    ws.merge_cells(start_row=r, start_column=1+c0, end_row=r+1, end_column=1+c0)
    _hdr(ws, r, 1+c0, "Week ending\n/ Region", C_NAVY)

    ws.merge_cells(start_row=r, start_column=2+c0, end_row=r+1, end_column=2+c0)
    _hdr(ws, r, 2+c0, "Share in\nGlobal Index", C_NAVY)

    ws.merge_cells(start_row=r, start_column=3+c0, end_row=r, end_column=5+c0)
    _hdr(ws, r, 3+c0, "Weekly Average Price", C_GREY, fg=C_BLACK)

    ws.merge_cells(start_row=r, start_column=6+c0, end_row=r+1, end_column=6+c0)
    _hdr(ws, r, 6+c0, "Index Value\n(Year 2000 = 100)", C_BLUE)

    ws.merge_cells(start_row=r, start_column=7+c0, end_row=r, end_column=9+c0)
    _hdr(ws, r, 7+c0, "Weekly Average Price versus", C_PALE_BG, fg=C_BLACK)

    # Sub-header row 2
    r += 1
    for c in (1+c0, 2+c0, 6+c0):
        ws.cell(row=r, column=c).border = _border()
    _hdr(ws, r, 3+c0, "cts/gal",               C_GREY,    fg=C_BLACK)
    _hdr(ws, r, 4+c0, "$/bbl",                 C_GREY,    fg=C_BLACK)
    _hdr(ws, r, 5+c0, "$/t",                   C_GREY,    fg=C_BLACK)
    _hdr(ws, r, 7+c0, "prior week's\naverage",  C_PALE_BG, fg=C_BLACK)
    _hdr(ws, r, 8+c0, "prior month's\naverage", C_PALE_BG, fg=C_BLACK)
    _hdr(ws, r, 9+c0, "prior year's\naverage",  C_PALE_BG, fg=C_BLACK)

    # Data rows
    for di, row in enumerate(rows):
        r += 1
        padded     = (row + [""] * 9)[:9]
        region     = padded[0]
        is_total   = "jet fuel" in region.lower() or di == 0
        is_special = "oil" in region.lower() or "crack" in region.lower()
        alt_bg     = C_ALT_ROW if (di % 2 == 1 and not is_total and not is_special) else None

        for ci_z, val in enumerate(padded):
            ci      = ci_z + 1 + c0
            bold    = is_total or is_special
            h_align = "left" if ci_z < 2 else "right"
            if ci_z == 5:
                bg = C_IDX_DATA
            elif ci_z >= 6:
                bg = C_PALE_BG
            else:
                bg = alt_bg
            _cell(ws, r, ci, val, bold=bold, bg=bg, h_align=h_align)

    return r


# ── Snapshot: Table 2 (cols 1-5) ─────────────────────────────────────────────
def _write_table2(ws, rows, start_row, col_offset=0):
    """
    Write Table 2. col_offset same as _write_table1.
    Returns last written row.
    """
    r  = start_row
    c0 = col_offset

    _hdr(ws, r, 1+c0, "Week ending",                       C_NAVY)
    _hdr(ws, r, 2+c0, "Index Value\n(Year 2000 = 100)",    C_BLUE)
    _hdr(ws, r, 3+c0, "Weekly Average Price\n$/bbl",       C_GREY, fg=C_BLACK)
    _hdr(ws, r, 4+c0, "Change vs\nprior week",              C_GREY, fg=C_BLACK)
    _hdr(ws, r, 5+c0, "Weekly Average\nCrack Spread $/bbl", C_GREY, fg=C_BLACK)

    for di, row in enumerate(rows):
        r += 1
        padded = (row + [""] * 5)[:5]
        alt_bg = C_ALT_ROW if di % 2 == 1 else None
        for ci_z, val in enumerate(padded):
            ci      = ci_z + 1 + c0
            h_align = "left" if ci_z == 0 else "right"
            bg      = C_IDX_DATA if ci_z == 1 else alt_bg
            _cell(ws, r, ci, val, bg=bg, h_align=h_align)

    return r


# ── Column sizing ─────────────────────────────────────────────────────────────
def _size_snapshot(ws):
    """Column widths for a snapshot sheet (no col A date column)."""
    for col, w in {1:26, 2:10, 3:11, 4:10, 5:12, 6:14, 7:12, 8:12, 9:12}.items():
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.row_dimensions[1].height = 38
    ws.row_dimensions[2].height = 28

def _size_consol_t1(ws):
    """Column widths for consolidated T1 (col A = date, then 9 data cols)."""
    ws.column_dimensions["A"].width = 18   # Extraction Date
    for col, w in {2:26, 3:10, 4:11, 5:10, 6:12, 7:14, 8:12, 9:12, 10:12}.items():
        ws.column_dimensions[get_column_letter(col)].width = w

def _size_consol_t2(ws):
    """Column widths for consolidated T2 (col A = date, then 5 data cols)."""
    ws.column_dimensions["A"].width = 18
    for col, w in {2:20, 3:14, 4:14, 5:13, 6:16}.items():
        ws.column_dimensions[get_column_letter(col)].width = w


# ── Consolidated sheet updater ────────────────────────────────────────────────
def _next_empty_row(ws) -> int:
    """Return the first fully empty row after all existing data."""
    return ws.max_row + 1 if ws.max_row and ws.max_row > 0 else 1


def _append_to_consolidated_t1(ws, rows, extraction_date: str):
    """
    Append a new Table-1 block to the consolidated sheet.
    Layout:
      - If sheet is empty: write col-A header then data block (col_offset=1).
      - If sheet has data: blank spacer row, then new data block.
    Col A header = "Extraction Date"; filled only on first data row of each block.
    """
    is_empty = ws.max_row <= 1 and not any(
        ws.cell(1, c).value for c in range(1, 12)
    )

    if is_empty:
        # Write "Extraction Date" column header
        _hdr(ws, 1, 1, "Extraction\nDate", C_NAVY)
        ws.row_dimensions[1].height = 38
        ws.row_dimensions[2].height = 28
        start = 1
    else:
        # Blank spacer row then start new block
        spacer = _next_empty_row(ws)
        ws.row_dimensions[spacer].height = 8
        start = spacer + 1

    # Write table headers (col_offset=1 shifts data to cols B-J)
    last = _write_table1(ws, rows, start_row=start, col_offset=1)

    # Merge col A across all data rows of this block (T1 has 2 header rows)
    first_data_row = start + 2
    if last >= first_data_row:
        ws.merge_cells(start_row=first_data_row, start_column=1,
                       end_row=last,             end_column=1)
    _cell(ws, first_data_row, 1, extraction_date,
          bold=True, bg=C_DATE_BG, h_align="center", wrap=True)

    _size_consol_t1(ws)


def _append_to_consolidated_t2(ws, rows, extraction_date: str):
    """Append a new Table-2 block to the consolidated sheet."""
    is_empty = ws.max_row <= 1 and not any(
        ws.cell(1, c).value for c in range(1, 7)
    )

    if is_empty:
        _hdr(ws, 1, 1, "Extraction\nDate", C_NAVY)
        ws.row_dimensions[1].height = 38
        start = 1
    else:
        spacer = _next_empty_row(ws)
        ws.row_dimensions[spacer].height = 8
        start = spacer + 1

    last = _write_table2(ws, rows, start_row=start, col_offset=1)

    # Merge col A across all data rows of this block (T2 has 1 header row)
    first_data_row = start + 1
    if last >= first_data_row:
        ws.merge_cells(start_row=first_data_row, start_column=1,
                       end_row=last,             end_column=1)
    _cell(ws, first_data_row, 1, extraction_date,
          bold=True, bg=C_DATE_BG, h_align="center", wrap=True)

    _size_consol_t2(ws)


# ── Duplicate detection ───────────────────────────────────────────────────────
def _data_hash(t1, t2) -> str:
    raw = json.dumps([t1, t2], ensure_ascii=False, sort_keys=True)
    return hashlib.sha256(raw.encode()).hexdigest()[:16]

def _last_hash(wb) -> str | None:
    """Read hash stored in a hidden cell on the last snapshot sheet."""
    # Snapshot sheets are all sheets except the two consolidated ones
    snapshots = [s for s in wb.sheetnames
                 if s not in (SHEET_CONSOL_T1, SHEET_CONSOL_T2)]
    if not snapshots:
        return None
    try:
        return wb[snapshots[-1]].cell(row=1000, column=1).value
    except Exception:
        return None

def _stash_hash(ws, h: str):
    c = ws.cell(row=1000, column=1, value=h)
    c.font = Font(color="FFFFFF", size=1)


# ── Public API ────────────────────────────────────────────────────────────────
def build_or_update(t1_rows: list[list[str]],
                    t2_rows: list[list[str]]) -> tuple[Path, bool, str]:
    """
    Create workbook (first call) or append a sheet (subsequent calls).

    Workbook always has:
      • "Consolidated T1"  at position 0
      • "Consolidated T2"  at position 1
      • Snapshot sheets    appended after

    Returns: (workbook_path, is_duplicate, sheet_name)
    """
    WORKBOOK_PATH.parent.mkdir(parents=True, exist_ok=True)

    current_hash  = _data_hash(t1_rows, t2_rows)
    sheet_name    = datetime.now().strftime("%Y-%m-%d %H.%M")
    extraction_dt = datetime.now().strftime("%Y-%m-%d %H:%M")

    # ── Load or create workbook ───────────────────────────────────────────────
    if WORKBOOK_PATH.exists():
        wb = load_workbook(WORKBOOK_PATH)

        # Duplicate check against last snapshot
        if _last_hash(wb) == current_hash:
            snapshots  = [s for s in wb.sheetnames
                          if s not in (SHEET_CONSOL_T1, SHEET_CONSOL_T2)]
            last_sheet = snapshots[-1] if snapshots else sheet_name
            wb.close()
            return WORKBOOK_PATH, True, last_sheet

    else:
        wb = Workbook()
        # Remove default empty sheet
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]

    # ── Ensure consolidated sheets exist at positions 0 and 1 ─────────────────
    if SHEET_CONSOL_T1 not in wb.sheetnames:
        wb.create_sheet(title=SHEET_CONSOL_T1, index=0)
    if SHEET_CONSOL_T2 not in wb.sheetnames:
        wb.create_sheet(title=SHEET_CONSOL_T2, index=1)

    # ── Append to consolidated sheets ─────────────────────────────────────────
    _append_to_consolidated_t1(wb[SHEET_CONSOL_T1], t1_rows, extraction_dt)
    _append_to_consolidated_t2(wb[SHEET_CONSOL_T2], t2_rows, extraction_dt)

    # ── Create snapshot sheet ─────────────────────────────────────────────────
    ws = wb.create_sheet(title=sheet_name)

    last_t1 = _write_table1(ws, t1_rows, start_row=1, col_offset=0)
    spacer  = last_t1 + 1
    ws.row_dimensions[spacer].height = 8
    _write_table2(ws, t2_rows, start_row=spacer + 1, col_offset=0)
    _size_snapshot(ws)
    _stash_hash(ws, current_hash)

    wb.save(WORKBOOK_PATH)
    wb.close()

    return WORKBOOK_PATH, False, sheet_name